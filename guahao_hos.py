# -*- coding:utf-8 -*-
import random
import socket
import time
import re
import urllib
import urllib2
import json
from openpyxl import Workbook
from openpyxl import load_workbook

socket.setdefaulttimeout(10)
f = open('hospitalId.txt','r')

def get_mid(w1,w2,text) :
    pat = re.compile(w1+'(.*?)'+w2,re.S)
    result = pat.findall(text)
    return result[0]

def get_mid_all(w1,w2,text) :
    pat = re.compile(w1+'(.*?)'+w2,re.S)
    result = pat.findall(text)
    return result

def get_mid_op(w1,w2,text,i) :
    pat = re.compile(w1+'(.*?)'+w2,re.S)
    result = pat.findall(text)
    return result[i]

def clear(s) :
    s = s.replace(' ','')
    s = s.replace('	','')
    s = s.replace('\n','')
    s = s.replace('\r','')
    return s

class hospital :
    def __init__(self) :
        name = 'unknown' #get
        location = 'unknown' 
        phone = '0'
        info  = 'unknown' #get
        comment_num = '0'
        yuyue_num = '0' #get
        rank = 'unknown' #get
        website = 'unknown'
        follow = '0'
        service = '0'
        keshi = []
        houzhen_time = '0'
        Keshi_string = '无'



hospital_list = []
        


def get_hos_info(hospital_id) :
    global hospital_list 
    hospital_list.append(hospital())
    i = len(hospital_list) - 1
    def connect() :
        url = 'http://www.guahao.com/hospital/' + str(hospital_id)
        request = urllib2.Request(url)
        response = urllib2.urlopen(request)
        return response.read()

    def connect_for_info() :
        url = 'http://www.guahao.com/hospital/introduction/' + str(hospital_id)
        request = urllib2.Request(url)
        response = urllib2.urlopen(request)
        return response.read()

    def get_intro() :
        try:
            text = connect_for_info()
        except:
            print 'connect to guahao.com fail!'
        w1 = '<div class="introduction-content">'
        w2 = '</div>'
        tmp = get_mid(w1,w2,text)
        hospital_list[i].info = clear(tmp)
        hospital_list[i].info = hospital_list[i].info.replace('<p>','')
        hospital_list[i].info = hospital_list[i].info.replace('</p>','')
        hospital_list[i].info = hospital_list[i].info.replace('&rdquo;','"')
        hospital_list[i].info = hospital_list[i].info.replace('&ldquo;','"')
        hospital_list[i].info = hospital_list[i].info.replace('<div>','"')
        hospital_list[i].info = hospital_list[i].info.replace('</div>','"')
        
    def get_info():
        try:
            text = connect()
        except:
            print 'connect to guahao.com fail'
        try:
            w1 = '<title>'
            w2 = '_微医'
            hospital_list[i].name = get_mid(w1,w2,text)
        except:
            print 'load name fail!'
        try:
            w1 = '<span>预约量</span>'
            w2 = '<span>' 
            tmp = get_mid(w1,w2,text)
            w1 = '<strong>'
            w2 = '</strong>'
            tmp = get_mid(w1,w2,tmp)
            w1 = '                    '
            w2 = '\n'
            hospital_list[i].yuyue_num = get_mid(w1,w2,tmp)
        except:
            print 'load yuyue_num fail!'
        try:
            w1 = '<div class="detail word-break">'
            w2 = '</h1>'
            tmp = get_mid(w1,w2,text)
            w1 = '<span>'
            w2 = '</span>'
            hospital_list[i].rank = clear(get_mid(w1,w2,tmp))
        except:
            print 'load rank fail!'
        try:
            w1 = '<span>患者评价</span>'
            w2 = '<span>'
            tmp = get_mid(w1,w2,text)
            tmp = get_mid('<strong>','</strong>',tmp)
            hospital_list[i].comment_num = clear(tmp)
        except:
            print 'load comment_num fail!'
        try:
            w1 = '<span>候诊时间</span>'
            w2 = '</p>'
            tmp = get_mid(w1,w2,text)
            tmp = clear(get_mid('<strong>','</strong>',tmp))
            hospital_list[i].houzhen_time =  tmp  
        except:
            print 'load houzhen_time fail!'
        try:
            w1 = '<span title="'
            w2 = '">'
            hospital_list[i].location = get_mid(w1,w2,text)
        except:
            print 'load location fail!'
        try:
            w1 = '<span>导医服务</span>'
            w2 = '</p>'
            tmp = get_mid(w1,w2,text)
            hospital_list[i].service = clear(get_mid('<strong>','</strong>',tmp))
        except:
            print 'load service fail!'
        try:
            w1 = '<b>电话：</b>'
            w2 = '</div>'
            tmp = get_mid(w1,w2,text)
            hospital_list[i].phone = clear(get_mid('<span>','</span>',tmp))
        except:
            print 'load phone fail!'
        try:
            w1 = '<b>官网：</b>'
            w2 = '</div>'
            tmp = get_mid(w1,w2,text)
            hospital_list[i].website = get_mid('<span>&nbsp;','</span>',tmp)
        except:
            hospital_list[i].website = '无'
        try:
            w1 = '<span class="mark-count">'
            w2 = '</span>'
            hospital_list[i].follow = get_mid(w1,w2,text)
        except:
            print 'load follow fail!'
        try:
            w1 = 'KSLB'
            w2 = '</em>'
            hospital_list[i].keshi = get_mid_all(w1,w2,text)
            for k in range(len(hospital_list[i].keshi)) :
                hospital_list[i].keshi[k] = hospital_list[i].keshi[k].replace('</a>','')
                hospital_list[i].keshi[k] = hospital_list[i].keshi[k].replace('<em>','')
                hospital_list[i].keshi[k] = hospital_list[i].keshi[k].replace('&nbsp;','')
                hospital_list[i].keshi[k] = hospital_list[i].keshi[k].replace('''')">''','')
                hospital_list[i].keshi[k] = clear(hospital_list[i].keshi[k])
            hospital_list[i].Keshi_string = '' 
            for k in hospital_list[i].keshi :
                hospital_list[i].Keshi_string += k 
                hospital_list[i].Keshi_string += ';'
        except:
            print 'load keshi fail!'

    get_info()
    get_intro()


def save_to_xlxs(i,x):
    global hospital_list
    wb = load_workbook('hospital2.xlsx')
    s = wb.active
    s.cell(row = i+x,column = 1,value = hospital_list[i-1].name)
    s.cell(row = i+x,column = 2,value = hospital_list[i-1].rank)
    s.cell(row = i+x,column = 3,value = hospital_list[i-1].location)
    s.cell(row = i+x,column = 4,value = hospital_list[i-1].phone)
    s.cell(row = i+x,column = 5,value = hospital_list[i-1].website)
    s.cell(row = i+x,column = 6,value = hospital_list[i-1].info)
    s.cell(row = i+x,column = 7,value = hospital_list[i-1].follow)
    s.cell(row = i+x,column = 8,value = hospital_list[i-1].yuyue_num)
    s.cell(row = i+x,column = 9,value = hospital_list[i-1].service)
    s.cell(row = i+x,column = 10,value = hospital_list[i-1].comment_num)
    s.cell(row = i+x,column = 11,value = hospital_list[i-1].houzhen_time)
    s.cell(row = i+x,column = 12,value = hospital_list[i-1].Keshi_string)
    wb.save('hospital2.xlsx')

def main() :
    i = 0
    x = input('fitst line:')
    y = input('last line:')
    for q in range(x-1) :
        f.readline()
    while 1 :
        i += 1
        if i+x-1 > y :
            break
        print 'getting Num.' + str(i+x-1) + ''' hospital's information'''
        hospital_id = f.readline().strip()
        if not hospital_id :
            break
        try:
            get_hos_info(hospital_id)
            save_to_xlxs(i,x)
        except:
            continue

if __name__ == "__main__" :
    main()

    
