# -*- coding:utf-8 -*-
import random
import socket
import time
import re
import urllib
import urllib2
import json
from openpyxl import Workbook
from openpyxl import load_workbook
#import xlrd
#import xlwt
#from xlutils.copy import copy

socket.setdefaulttimeout(10)

doctor_list =[]

t = input('waiting time:')
et = input('error waiting time:')

class doctor :
    def __init__(self) :
        expert_id = ''
        name = ''
        skill = ''
        info = '暂无'
        Yuyue = '' 
        Wenzhen = ''
        hospital = ''
        dept = ''
        dept2 = ''
        cid = ''
        pid = ''
        rank = ''
        mark_count = '' 
        keys = []
        rate = ''
        share_num = ''
        Tuwen = '0'
        Dianhua = '0'
        Shipin = '0'
        Guahao = '0'
        team = ''
        WenzhenNum = ''
        clinicType = ''
        serveType =''

        

def get_mid(w1,w2,text) :
    pat = re.compile(w1+'(.*?)'+w2,re.S)
    result = pat.findall(text)
    return result[0];

def get_mid_op(w1,w2,text,i) :
    pat = re.compile(w1+'(.*?)'+w2,re.S)
    result = pat.findall(text)
    return result[i];

def get_mid_all(w1,w2,text) :
    pat = re.compile(w1+'(.*?)'+w2,re.S)
    result = pat.findall(text)
    return result

def get_doc_info(expert_id,x) :
    def get_info_needed(text,i) :
        try:
            w1 = '<div class="detail word-break">'
            w2 = '</h1>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].name = get_mid('<strong>','</strong>',tmp)
        except:
            doctor_list[i].name = '   '
        try:
            doctor_list[i].rank = get_mid('<span> ','</span>',tmp)
        except:
            doctor_list[i].rank = '暂无'
        try:
            w1 = '<a title="'
            w2 = '"'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].hospital = tmp
            doctor_list[i].hospital2 = get_mid_op(w1,w2,text,1)
        except:
            print 'load hospital fail!'
        try:
            w1 = '<a href="http://www.guahao.com/department'
            w2 = '</a>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].dept = get_mid('\n                                        ','\n',tmp)
        except:
            doctor_list[i].dept = '未知'
        try:
            w1 = '<a href="http://www.guahao.com/department'
            w2 = '</a>'
            tmp = get_mid_op(w1,w2,text,1)
            doctor_list[i].dept2 = get_mid('\n                                        ','\n',tmp)
        except:
            doctor_list[i].dept2 = '  '
        try:
            w1 = '<span class="mark-count">'
            w2 = '</span>'
            doctor_list[i].mark_count = get_mid(w1,w2,text)
        except:
            doctor_list[i].mark_count = '0' 
            print 'load mark_count fail!'
        try:
            w1 = 'data-description="'
            w2 = '">'
            doctor_list[i].info = get_mid_op(w1,w2,text,1)
        except:
            try:
                w1 = '<b>简介：</b>'
                w2 = '</div>'
                tmp = get_mid(w1,w2,text)
                doctor_list[i].info = get_mid('<span>','</span>',tmp)
            except:
                doctor_list[i].info = '暂无'
        try:
            w1 = '<b>擅长：</b>'
            w2 = '</div>'
            tmp = get_mid(w1,w2,text)
            w1 = '<span>'
            w2 = '</span>'
            doctor_list[i].skill = get_mid(w1,w2,tmp)
        except:
            doctor_list[i].skill = '暂无相关信息'
        try:
            w1 = '<div class="keys">'
            w2 = '</div>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].keys = get_mid_all('title="','">',tmp)
        except:
            doctor_list[i].keys = '无'
            print 'laod keys fail!'
        try:
            w1 = '<span>预约量</span>'
            w2 = '<span>问诊量</span>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].Yuyue = get_mid('<strong>','</strong>',tmp)
        except:
            doctor_list[i].Yuyue = '暂无'
            print 'load Yuyue fail!'
        try:
            w2 = '</div>'
            w1 = '<span>问诊量</span>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].Wenzhen = get_mid('<strong>','</strong>',tmp)
        except:
            doctor_list[i].Wenzhen = '暂无'
            print 'load Wenzhen fail!'
        try:
            w1 = '<p class="light"></p>'
            w2 = '</a>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].rate = get_mid('<strong>','</strong>',tmp)
            if doctor_list[i].rate == '.0' :
                doctor_list[i].rate = '暂无'
        except:
            doctor_list[i].rate = '暂无'
            print 'load rate fail!'
        try :
            w1 = '<h3>患者评价</h3>'
            w2 = '</div>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].share_num = get_mid('<strong>','</strong>',tmp)
        except :
            doctor_list[i].share_num = '0'
        try :
            w1 = '<div class="group-show-new" title="'
            w2 = '">'
            doctor_list[i].team = get_mid(w1,w2,text)
        except:
            doctor_list[i].team = ''
        try :
            w1 = '<h4>已累计为<strong>'
            w2 = '</strong>人次提供图文问诊'
            doctor_list[i].WenzhenNum = get_mid(w1,w2,text)
        except:
            doctor_list[i].WenzhenNum = '0'
        try:
            w1 = '<h4>图文问诊</h4>'
            w2 = '<span>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].Tuwen = get_mid('<strong>','</strong>',tmp)
        except:
            doctor_list[i].Tuwen = '暂未开通功能'
        try:
            w1 = '<h4>电话问诊</h4>'
            w2 = '<span>'
            tmp = get_mid(w1,w2,text)
            doctor_list[i].Dianhua = get_mid('<strong>','</strong>',tmp)
        except:
            doctor_list[i].Dianhua = '暂未开通功能'
        try:
            w1 = '<h4>视频问诊</h4>'
            w2 = '</span>'
            tmp = get_mid(w1,w2,text)
            if tmp.find('暂未开通功能') != -1 :
                doctor_list[i].Shipin = '暂未开通功能'
            else :
                doctor_list[i].Shipin = '已开通'
        except:
            doctor_list[i].Shipin = ' '
            
            
    def get_paiban(text,i) :
        try :
            w1 = '"clinicType":"'
            w2 = '",'
            doctor_list[i].clinicType = get_mid(w1,w2,text)
        except :
            doctor_list[i].clinicType = '暂无排班消息'
        try :
            w1 = '"price":'
            w2 = ',"'
            doctor_list[i].price = get_mid(w1,w2,text)
        except:
            doctor_list[i].price = '暂无排班消息'


    global doctor_list
    global t
    global et
    enable_proxy = True
    doctor_list.append(doctor())
    i = len(doctor_list) - 1
    doctor_list[i].expert_id = expert_id
    url = 'http://www.guahao.com/expert'
    geturl = url + "/" + expert_id
    purl = 'http://www.guahao.com/expert/new/shiftcase/?expertId=' + expert_id
    record = 0
    while(1):
        try:
            request = urllib2.Request(geturl)
            response = urllib2.urlopen(request)
            time.sleep(t/2)
            prequest = urllib2.Request(purl)
            presponse = urllib2.urlopen(prequest)
            break
        except:
            record += 1
            if record == 5 :
                print 'fail to get this data after trying 5 times!'
                break
            print "503 error,try again in 6 seconds"
            time.sleep(et)
            continue
    record = 0
    
    try:
        get_info_needed(response.read(),i)
        get_paiban(presponse.read(),i)
        if doctor_list[i].Tuwen != '暂未开通功能' :
            if doctor_list[i].Dianhua != '暂未开通功能' :
                if doctor_list[i].Shipin != '暂未开通功能' :
                    if doctor_list[i].clinicType != '暂无排班消息' :
                        doctor_list[i].serveType = '预约挂号、图文、电话、视频'
                    else:
                        doctor_list[i].serveType = '图文、电话、视频'
                else :
                    doctor_list[i].serveType = '图文、电话'
            else:
                doctor_list[i].serveType = '图文'
        else :
            doctor_list[i].serveType = '无'
    except:
        print 'something wrong'
    time.sleep(t - t/2)
    

def copy_to_xls(q,z,x) :
    global doctor_list
    wb = load_workbook('test.xlsx')
    s = wb.active
    record = 0
    while 1:
        try:
            print 'saving the datas to excel!do not close the app.'
            for i in range(z):
                try:
                    s.cell(row = q-i+x,column = 1,value = doctor_list[q-i-1].name)
                    s.cell(row = q-i+x,column = 2,value = doctor_list[q-i-1].mark_count)
                    s.cell(row = q-i+x,column = 3,value = doctor_list[q-i-1].team)
                    s.cell(row = q-i+x,column = 6,value = doctor_list[q-i-1].rank)
                    s.cell(row = q-i+x,column = 5,value = doctor_list[q-i-1].dept + '/' + doctor_list[q-i-1].dept2)
                    s.cell(row = q-i+x,column = 6,value = doctor_list[q-i-1].skill)
                    s.cell(row = q-i+x,column = 7,value = doctor_list[q-i-1].info)
                    s.cell(row = q-i+x,column = 8,value = doctor_list[q-i-1].serveType)
                    s.cell(row = q-i+x,column = 9,value = doctor_list[q-i-1].rate)
                    s.cell(row = q-i+x,column = 10,value = doctor_list[q-i-1].Yuyue)
                    s.cell(row = q-i+x,column = 11,value = doctor_list[q-i-1].Wenzhen)
                    s.cell(row = q-i+x,column = 12,value = doctor_list[q-i-1].clinicType)
                    s.cell(row = q-i+x,column = 13,value = doctor_list[q-i-1].price)
                    s.cell(row = q-i+x,column = 14,value = doctor_list[q-i-1].Tuwen)
                    s.cell(row = q-i+x,column = 15,value = doctor_list[q-i-1].Dianhua)
                    s.cell(row = q-i+x,column = 16,value = doctor_list[q-i-1].Shipin)
                    s.cell(row = q-i+x,column = 17,value = doctor_list[q-i-1].share_num)
                    s.cell(row = q-i+x,column = 18,value = doctor_list[q-i-1].WenzhenNum)
                    s.cell(row = q-i+x,column = 19,value = doctor_list[q-i-1].expert_id)
                except:
                    print 'write ' + doctor_list[q-i-1].expert_id + ' to excel fail!having saved the id to error.txt.'
                    err = file('error.txt','a')
                    err.write('\n' + expert_id + ' ' + 'Num.' + str(i+x))
                    err.close()
                    continue
            wb.save('test.xlsx')
            print 'save completed,you can exit the app now.'
        except:
            record += 1
            if record == 5:
                print '''having tried 5 times but still can't save the data!'''
                break
            print 'save to test.xlsx fail! trying the ' + str(record) + ' times now'
            continue
        break


def backup(name) :
    wb = load_workbook('test.xlsx')
    wb.save(name)

def main() :
    f = open('expertId.txt','r')
    i = 0
    x = input('first line:')
    y = input('last line:')
    z = input('save frequency:')

    for q in range(x-1) :
        f.readline()
    while 1 :
        i += 1
        print 'getting Num.' + str(i+x-1) + ''' except's information'''
        expert_id = f.readline().strip() 
        if not expert_id :
            break
        get_doc_info(expert_id,x)
        if i % z == 0 :
            copy_to_xls(i,z,x)
        try:
            if i%100 == 0 :
                backup('./backup/backup_for_100.xlsx')
            elif i%500 == 0 :
                backup('./backup/backup_for_500.xlsx')
            elif i%1000 == 0 :
                backup('./backup/backup_for_1000.xlsx')
        except:
            print 'backup fail!'
        if i == (y-x+1) :
            break

if __name__ == "__main__":
    main()

